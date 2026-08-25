#!/usr/bin/env python3
"""Actualiza la hoja ``Lista de ramales`` de ramales.xlsx.

SOFSE y Cuándo SUBO no exponen GraphQL. El catálogo se obtiene desde sus
endpoints REST de listado y se cachea en el propio configurador para que el
actualizador diario no tenga que redescubrir cientos de agencies.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import logging
import os
import re
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from scripts.update_schedules import (
        ARGENTINA_TZ,
        CUANDO_SUBO_API_KEY,
        CUANDO_SUBO_BASE_URL,
        DEFAULT_BRANCHES,
        JsonHttpClient,
        SofseProvider,
        SourceUnavailable,
        normalize_text,
    )
except ImportError:
    from update_schedules import (  # type: ignore[no-redef]
        ARGENTINA_TZ,
        CUANDO_SUBO_API_KEY,
        CUANDO_SUBO_BASE_URL,
        DEFAULT_BRANCHES,
        JsonHttpClient,
        SofseProvider,
        SourceUnavailable,
        normalize_text,
    )

CATALOG_HEADERS = [
    "Tipo",
    "Proveedor",
    "ID API",
    "Linea / ramal",
    "Nombre API",
    "Empresa / agencia",
    "Descripcion",
]
CATALOG_SHEET = "Lista de ramales"
CATALOG_DATE_CELL = "I2"
DEFAULT_MAX_AGE_DAYS = 7
MAX_AGENCY_WORKERS = 3

logger = logging.getLogger("solaris.catalog")


def _collection(response: Any) -> list[dict[str, Any]]:
    if isinstance(response, list):
        return response
    if not isinstance(response, dict):
        return []
    data = response.get("data")
    if isinstance(data, dict) and isinstance(data.get("list"), list):
        return data["list"]
    for key in ("results", "resultado", "list"):
        if isinstance(response.get(key), list):
            return response[key]
    return []


def _agency_routes(agency_id: str) -> tuple[str, str, list[dict[str, Any]]]:
    http = JsonHttpClient()
    query = f"key={CUANDO_SUBO_API_KEY}"
    response = http.request(
        f"{CUANDO_SUBO_BASE_URL}/routes-for-agency/{agency_id}.json?{query}"
    )
    if response.get("code") != 200:
        raise SourceUnavailable(
            f"agency {agency_id}: {response.get('text') or response.get('code')}"
        )
    references = response.get("data", {}).get("references", {})
    agency = next(
        (
            item for item in references.get("agencies", [])
            if str(item.get("id")) == agency_id
        ),
        {},
    )
    return agency_id, str(agency.get("name") or agency_id), _collection(response)


def _bus_catalog(http: JsonHttpClient) -> list[dict[str, str]]:
    response = http.request(
        f"{CUANDO_SUBO_BASE_URL}/agencies-with-coverage.json?key={CUANDO_SUBO_API_KEY}"
    )
    agency_ids = sorted({
        str(item.get("agencyId"))
        for item in _collection(response)
        if item.get("agencyId") is not None
    })
    if not agency_ids:
        raise SourceUnavailable("Cuándo SUBO no devolvió agencies")

    routes: list[tuple[str, str, dict[str, Any]]] = []
    failed_ids: list[str] = []
    with ThreadPoolExecutor(max_workers=MAX_AGENCY_WORKERS) as executor:
        futures = {executor.submit(_agency_routes, agency_id): agency_id for agency_id in agency_ids}
        for future in as_completed(futures):
            agency_id = futures[future]
            try:
                _, agency_name, agency_routes = future.result()
            except Exception as exc:  # cada request ya aplica reintentos acotados
                logger.debug("Primera pasada fallida para agency %s: %s", agency_id, exc)
                failed_ids.append(agency_id)
                continue
            routes.extend(
                (agency_id, agency_name, route) for route in agency_routes
            )
    failures: list[str] = []
    for agency_id in failed_ids:
        time.sleep(0.2)
        try:
            _, agency_name, agency_routes = _agency_routes(agency_id)
        except Exception as exc:
            failures.append(f"{agency_id}: {exc}")
            continue
        routes.extend((agency_id, agency_name, route) for route in agency_routes)

    if failures:
        sample = "; ".join(failures[:5])
        raise SourceUnavailable(
            f"Catálogo incompleto: fallaron {len(failures)} de {len(agency_ids)} agencies ({sample})"
        )

    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    for agency_id, agency_name, route in routes:
        description = str(route.get("description") or "").strip()
        base, separator, direction = description.rpartition(":")
        if not separator:
            base, direction = description, ""
        line = str(route.get("shortName") or route.get("longName") or "").strip()
        key = (agency_id, normalize_text(line), normalize_text(base))
        group = grouped.setdefault(key, {
            "agency": agency_name,
            "line": line,
            "base": base,
            "ids": [],
            "directions": [],
        })
        route_id = str(route.get("id") or "").strip()
        if route_id and route_id not in group["ids"]:
            group["ids"].append(route_id)
        if direction and direction.strip() not in group["directions"]:
            group["directions"].append(direction.strip())

    rows = []
    for group in grouped.values():
        ids = sorted(group["ids"], key=lambda value: int(value.rsplit("_", 1)[-1]))
        name = " - ".join(value for value in (group["line"], group["base"]) if value)
        rows.append({
            "Tipo": "Colectivo",
            "Proveedor": "Cuando SUBO",
            "ID API": " / ".join(ids),
            "Linea / ramal": group["line"],
            "Nombre API": name,
            "Empresa / agencia": group["agency"],
            "Descripcion": " / ".join(group["directions"]),
        })
    return rows


def _train_catalog(http: JsonHttpClient) -> list[dict[str, str]]:
    provider = SofseProvider(http)
    managements = provider._response_items(
        provider.get("/infraestructura/gerencias", {"idEmpresa": 1})
    )
    rows: list[dict[str, str]] = []
    for management in managements:
        management_id = management.get("id")
        if management_id is None:
            continue
        branches = provider._response_items(
            provider.get("/infraestructura/ramales", {"idGerencia": management_id})
        )
        for branch in branches:
            if branch.get("id") is None:
                continue
            rows.append({
                "Tipo": "Tren",
                "Proveedor": "SOFSE",
                "ID API": str(branch["id"]),
                "Linea / ramal": str(management.get("nombre") or ""),
                "Nombre API": str(branch.get("nombre") or branch["id"]),
                "Empresa / agencia": "Trenes Argentinos",
                "Descripcion": str(branch.get("nombre") or ""),
            })
    if not rows:
        raise SourceUnavailable("SOFSE no devolvió ramales")
    return rows


def build_catalog() -> list[dict[str, str]]:
    http = JsonHttpClient()
    rows = _train_catalog(http) + _bus_catalog(http)
    rows.sort(key=lambda row: (
        0 if row["Tipo"] == "Tren" else 1,
        normalize_text(row["Linea / ramal"]),
        normalize_text(row["Nombre API"]),
    ))
    return rows


def catalog_age_days(path: Path, today: dt.date) -> int | None:
    if not path.exists():
        return None
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if CATALOG_SHEET not in workbook.sheetnames:
            return None
        value = workbook[CATALOG_SHEET][CATALOG_DATE_CELL].value
        if isinstance(value, dt.datetime):
            value = value.date()
        if isinstance(value, str):
            value = dt.date.fromisoformat(value[:10])
        return (today - value).days if isinstance(value, dt.date) else None
    except (TypeError, ValueError):
        return None
    finally:
        workbook.close()


def write_catalog(path: Path, rows: list[dict[str, str]], today: dt.date) -> None:
    workbook = load_workbook(path)
    try:
        if CATALOG_SHEET in workbook.sheetnames:
            del workbook[CATALOG_SHEET]
        sheet = workbook.create_sheet(CATALOG_SHEET)
        sheet.append(CATALOG_HEADERS)
        for row in rows:
            sheet.append([row.get(header, "") for header in CATALOG_HEADERS])

        header_fill = PatternFill("solid", fgColor="17365D")
        for cell in sheet[1]:
            if cell.column <= len(CATALOG_HEADERS):
                cell.fill = header_fill
                cell.font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows(min_row=2, max_col=len(CATALOG_HEADERS)):
            for cell in row:
                cell.font = Font(name="Aptos", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        widths = [13, 16, 25, 20, 60, 42, 70]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:G{len(rows) + 1}"
        table = Table(displayName="CatalogoRamales", ref=f"A1:G{len(rows) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        sheet["I1"] = "Actualizado"
        sheet[CATALOG_DATE_CELL] = today
        sheet[CATALOG_DATE_CELL].number_format = "yyyy-mm-dd"
        sheet["I3"] = "Ramales"
        sheet["I4"] = len(rows)
        sheet.column_dimensions["I"].width = 15

        candidate: Path | None = None
        try:
            with tempfile.NamedTemporaryFile(
                prefix=".solaris-ramales-", suffix=".xlsx", dir=path.parent, delete=False,
            ) as temp_file:
                candidate = Path(temp_file.name)
            workbook.save(candidate)
            check = load_workbook(candidate, read_only=True, data_only=True)
            try:
                if CATALOG_SHEET not in check.sheetnames or check[CATALOG_SHEET].max_row != len(rows) + 1:
                    raise RuntimeError("La verificación del catálogo XLSX falló")
            finally:
                check.close()
            os.replace(candidate, path)
            candidate = None
        finally:
            if candidate is not None:
                candidate.unlink(missing_ok=True)
    finally:
        workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--branches", type=Path, default=DEFAULT_BRANCHES)
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--max-age-days", type=int, default=DEFAULT_MAX_AGE_DAYS)
    parser.add_argument("--output-json", type=Path)
    parser.add_argument("--verbose", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )
    today = dt.datetime.now(ARGENTINA_TZ).date()
    age = catalog_age_days(args.branches, today)
    if not args.force and not args.output_json and age is not None and age < args.max_age_days:
        logger.info("Catálogo vigente (%s días); no se consulta la API", age)
        return 0

    rows = build_catalog()
    if args.output_json:
        args.output_json.parent.mkdir(parents=True, exist_ok=True)
        args.output_json.write_text(
            json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8",
        )
        logger.info("Catálogo exportado a %s (%s ramales)", args.output_json, len(rows))
        return 0
    if not args.branches.exists():
        logger.error("No existe %s; cree primero el configurador", args.branches)
        return 2
    write_catalog(args.branches, rows, today)
    logger.info("Catálogo actualizado: %s ramales", len(rows))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
