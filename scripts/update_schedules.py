#!/usr/bin/env python3
"""Actualiza snapshots XLSX desde las APIs usadas por las apps oficiales.

La regla central es conservadora: una fuente vacia, parcial o inaccesible nunca
reemplaza un libro existente. Cada libro se construye en un temporal, se valida
con el mismo parser que alimenta SQLite y recien entonces se reemplaza de forma
atomica.
"""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import difflib
import json
import logging
import os
import re
import ssl
import tempfile
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CONFIG = PROJECT_ROOT / "config" / "schedule_sources.json"
DEFAULT_BRANCHES = PROJECT_ROOT / "ramales.xlsx"
ARGENTINA_TZ = ZoneInfo("America/Argentina/Buenos_Aires")
SOFSE_BASE_URL = "https://api-servicios.sofse.gob.ar/v1"
CUANDO_SUBO_BASE_URL = (
    "https://cuandosubo.sube.gob.ar/onebusaway-api-webapp/api/where"
)
CUANDO_SUBO_API_KEY = "web"
DEFAULT_TIMEOUT = 15
HTTP_RETRIES = 2
MAX_AUTODISCOVERED_BUS_STOPS = 12

logger = logging.getLogger("solaris.update")


class ConfigurationError(Exception):
    """La configuracion local es invalida y la ejecucion debe fallar."""


class SourceUnavailable(Exception):
    """La fuente remota no pudo producir un snapshot seguro."""


class IncompleteSchedule(SourceUnavailable):
    """La respuesta existe, pero no alcanza el umbral para reemplazar datos."""


@dataclass(frozen=True)
class ScheduleSnapshot:
    stations: list[str]
    formations: list[str]
    matrix: list[list[int | None]]
    source_date: dt.date


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text.casefold()).strip()


def slugify(value: Any) -> str:
    normalized = normalize_text(value)
    return re.sub(r"[^a-z0-9]+", "-", normalized).strip("-")


def compact_suffix(value: Any) -> str:
    words = re.findall(r"[a-z0-9]+", normalize_text(value))
    return "".join(word.capitalize() for word in words) or "Destino"


def folder_component(value: Any) -> str:
    words = re.findall(r"[a-z0-9]+", normalize_text(value))
    return "-".join(word.capitalize() for word in words) or "Ramal"


def _configured_route_names(route: dict[str, Any]) -> set[str]:
    values = {
        route.get("id"),
        route.get("route"),
        route.get("branch"),
        f"{route.get('branch', '')} {route.get('route', '')}",
        *(route.get("selectors") or []),
    }
    return {normalize_text(value) for value in values if value}


def load_branch_selections(path: Path) -> dict[str, list[str]]:
    """Lee únicamente las dos columnas editables de ramales.xlsx."""
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, ValueError) as exc:
        raise ConfigurationError(f"No se pudo leer el configurador {path}: {exc}") from exc
    try:
        if "Configurador" not in workbook.sheetnames:
            raise ConfigurationError(
                f"{path} no contiene la hoja Configurador"
            )
        sheet = workbook["Configurador"]
        headers = {
            normalize_text(sheet.cell(row=1, column=column).value): column
            for column in range(1, 3)
        }
        if "tren" not in headers or "colectivos" not in headers:
            raise ConfigurationError(
                "La hoja Configurador debe tener las columnas Tren y Colectivos"
            )
        result = {"Tren": [], "Colectivo": []}
        for type_name, header in (("Tren", "tren"), ("Colectivo", "colectivos")):
            seen: set[str] = set()
            column = headers[header]
            for row_values in sheet.iter_rows(
                min_row=2, min_col=column, max_col=column, values_only=True,
            ):
                value = str(row_values[0] or "").strip()
                key = normalize_text(value)
                if value and key not in seen:
                    result[type_name].append(value)
                    seen.add(key)
        if not any(result.values()):
            raise ConfigurationError("ramales.xlsx no contiene ramales para procesar")
        return result
    finally:
        workbook.close()


def load_branch_catalog(path: Path) -> list[dict[str, str]]:
    """Carga el catálogo técnico cacheado en la segunda hoja."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Lista de ramales" not in workbook.sheetnames:
            return []
        sheet = workbook["Lista de ramales"]
        headers = [
            str(sheet.cell(row=1, column=column).value or "").strip()
            for column in range(1, 8)
        ]
        rows: list[dict[str, str]] = []
        for values in sheet.iter_rows(min_row=2, min_col=1, max_col=7, values_only=True):
            row = {
                header: str(value or "").strip()
                for header, value in zip(headers, values)
                if header
            }
            if row.get("Tipo") and row.get("ID API"):
                rows.append(row)
        return rows
    finally:
        workbook.close()


def _catalog_match(
    selection: str, type_name: str, catalog: list[dict[str, str]],
) -> dict[str, str] | None:
    wanted = normalize_text(selection)
    wanted_key = slugify(selection)
    candidates: list[tuple[float, dict[str, str]]] = []
    requested_ids = set(re.findall(r"\d+_\d+|(?<!\d)\d+(?!\d)", selection))
    for row in catalog:
        if normalize_text(row.get("Tipo")) != normalize_text(type_name):
            continue
        row_ids = set(re.findall(r"\d+_\d+|(?<!\d)\d+(?!\d)", row.get("ID API", "")))
        fields = [
            row.get("Nombre API", ""),
            row.get("Linea / ramal", ""),
            row.get("Descripcion", ""),
            row.get("ID API", ""),
        ]
        normalized_fields = [normalize_text(field) for field in fields if field]
        if (
            any(wanted_key == slugify(field) for field in fields if field)
            or wanted in normalized_fields
            or (requested_ids and requested_ids <= row_ids)
        ):
            return row
        haystack = " ".join(normalized_fields)
        ratio = difflib.SequenceMatcher(None, wanted, haystack).ratio()
        wanted_tokens = set(wanted.split())
        token_score = len(wanted_tokens & set(haystack.split())) / max(1, len(wanted_tokens))
        candidates.append((ratio * 0.45 + token_score * 0.55, row))
    candidates.sort(key=lambda pair: pair[0], reverse=True)
    if not candidates or candidates[0][0] < 0.58:
        return None
    if len(candidates) > 1 and candidates[0][0] - candidates[1][0] < 0.04:
        return None
    return candidates[0][1]


def next_weekday(today: dt.date, weekday: int) -> dt.date:
    if weekday < 0 or weekday > 6:
        raise ConfigurationError(f"Dia de semana invalido: {weekday}")
    return today + dt.timedelta(days=(weekday - today.weekday()) % 7)


def parse_iso_timestamp(value: str) -> dt.datetime:
    if not value:
        raise ValueError("Timestamp vacio")
    parsed = dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=dt.timezone.utc)
    return parsed.astimezone(ARGENTINA_TZ)


def time_from_iso(value: str | None, source_date: dt.date) -> int | None:
    if not value:
        return None
    parsed = parse_iso_timestamp(value)
    day_offset = (parsed.date() - source_date).days
    if day_offset < 0 or day_offset > 1:
        return None
    return day_offset * 1440 + parsed.hour * 60 + parsed.minute


def time_from_epoch_millis(value: int | float | None, source_date: dt.date) -> int | None:
    if value is None:
        return None
    parsed = dt.datetime.fromtimestamp(float(value) / 1000, tz=ARGENTINA_TZ)
    day_offset = (parsed.date() - source_date).days
    if day_offset < 0 or day_offset > 1:
        return None
    return day_offset * 1440 + parsed.hour * 60 + parsed.minute


def minute_value(value: int | None) -> int | None:
    return value


def validate_snapshot(snapshot: ScheduleSnapshot, allow_short_turns: bool = False) -> None:
    if len(snapshot.stations) < 2:
        raise IncompleteSchedule("Se necesitan al menos dos estaciones")
    if not snapshot.formations or not snapshot.matrix:
        raise IncompleteSchedule("La fuente no devolvio servicios")
    if len(snapshot.formations) != len(snapshot.matrix):
        raise IncompleteSchedule("Formaciones y matriz tienen distinta longitud")
    if len(set(snapshot.formations)) != len(snapshot.formations):
        raise IncompleteSchedule("La fuente devolvio formaciones duplicadas")

    complete_destinations = 0
    for row in snapshot.matrix:
        if len(row) != len(snapshot.stations):
            raise IncompleteSchedule("Fila con cantidad de estaciones invalida")
        present = [value for value in row if value is not None]
        if len(present) < 2:
            raise IncompleteSchedule("Servicio con menos de dos horarios")
        if any(current < previous for previous, current in zip(present, present[1:])):
            raise IncompleteSchedule("Servicio con horarios decrecientes")
        if row[-1] is not None:
            complete_destinations += 1

    if not allow_short_turns:
        coverage = complete_destinations / len(snapshot.matrix)
        if coverage < 0.8:
            raise IncompleteSchedule(
                f"Cobertura del destino insuficiente: {coverage:.0%} (minimo 80%)"
            )


class JsonHttpClient:
    def __init__(self, timeout: int = DEFAULT_TIMEOUT):
        self.timeout = timeout
        self.ssl_context = ssl.create_default_context()

    def request(
        self,
        url: str,
        *,
        method: str = "GET",
        headers: dict[str, str] | None = None,
        body: dict[str, Any] | None = None,
    ) -> Any:
        data = None if body is None else json.dumps(body).encode("utf-8")
        request_headers = {
            "Accept": "application/json",
            "User-Agent": "SolarisPKN-Transport/1.0",
            **(headers or {}),
        }
        if data is not None:
            request_headers["Content-Type"] = "application/json"

        last_error: Exception | None = None
        for attempt in range(HTTP_RETRIES + 1):
            try:
                request = urllib.request.Request(
                    url, data=data, headers=request_headers, method=method,
                )
                with urllib.request.urlopen(
                    request, timeout=self.timeout, context=self.ssl_context,
                ) as response:
                    payload = response.read().decode("utf-8-sig")
                return json.loads(payload)
            except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
                last_error = exc
                if attempt < HTTP_RETRIES:
                    time.sleep(0.5 * (attempt + 1))

        raise SourceUnavailable(f"No se pudo consultar {url}: {last_error}")


def _replace_characters(value: str, replacements: dict[str, str]) -> str:
    return "".join(replacements.get(character, character) for character in value)


def build_sofse_credentials(now: dt.datetime) -> dict[str, str]:
    local = now.astimezone(ARGENTINA_TZ)
    compact_date = local.strftime("%Y%m%d")
    username = base64.b64encode(f"{compact_date}sofse".encode()).decode()

    first = base64.b64encode(username.encode()).decode()
    first = _replace_characters(first, {
        "a": "#t", "e": "#x", "i": "#f", "o": "#l", "u": "#7", "=": "#g",
    })[::-1]
    second = base64.b64encode(first.encode()).decode()
    second = _replace_characters(second, {
        "a": "#j", "e": "#p", "i": "#w", "o": "#8", "u": "#0", "=": "#v",
    })[::-1]

    return {
        "username": username,
        "password": urllib.parse.quote(second, safe="-_.!~*'()"),
    }


class SofseProvider:
    def __init__(self, http: JsonHttpClient):
        self.http = http
        self.token: str | None = None
        self.token_date: dt.date | None = None

    def authenticate(self) -> None:
        now = dt.datetime.now(ARGENTINA_TZ)
        response = self.http.request(
            f"{SOFSE_BASE_URL}/auth/authorize",
            method="POST",
            body=build_sofse_credentials(now),
        )
        token = response if isinstance(response, str) else (
            response.get("token") or response.get("accessToken")
            or response.get("access_token")
        )
        if not token:
            raise SourceUnavailable("SOFSE no devolvio un token reconocible")
        self.token = token
        self.token_date = now.date()

    def get(self, path: str, query: dict[str, Any]) -> Any:
        today = dt.datetime.now(ARGENTINA_TZ).date()
        if not self.token or self.token_date != today:
            self.authenticate()
        encoded = urllib.parse.urlencode({
            key: value for key, value in query.items() if value is not None
        })
        return self.http.request(
            f"{SOFSE_BASE_URL}{path}?{encoded}",
            headers={"Authorization": self.token or ""},
        )

    @staticmethod
    def _station_time(station: dict[str, Any], source_date: dt.date) -> int | None:
        departure = station.get("salida", {}).get("programada")
        arrival = station.get("llegada", {}).get("programada")
        return time_from_iso(departure or arrival, source_date)

    def snapshot(
        self, route: dict[str, Any], direction: dict[str, Any], source_date: dt.date,
    ) -> ScheduleSnapshot:
        query_station_ids = direction.get("query_station_ids") or [
            direction["query_station_id"]
        ]
        items: list[dict[str, Any]] = []
        for query_station_id in query_station_ids:
            response = self.get(
                f"/arribos/estacion/{query_station_id}",
                {
                    "hasta": direction.get("destination_id"),
                    "fecha": source_date.isoformat(),
                    "hora": "00:00",
                    "cantidad": 100,
                    "paraApp": "true",
                    "ramal": direction["branch_id"],
                    "sentido": direction["direction"],
                },
            )
            items.extend(
                response if isinstance(response, list) else response.get("results", [])
            )
        station_ids = [str(station["id"]) for station in direction["stations"]]
        rows: dict[str, tuple[str, list[dt.time | None]]] = {}

        for item in items:
            service = item.get("servicio", item)
            service_date = service.get("fecha")
            if service_date and parse_iso_timestamp(service_date).date() != source_date:
                continue
            if int(service.get("sentido", direction["direction"])) != int(direction["direction"]):
                continue

            by_station = {
                str(station.get("idElemento")): station
                for station in service.get("estaciones", [])
                if station.get("idElemento") is not None
            }
            matrix_row = [
                self._station_time(by_station[station_id], source_date)
                if station_id in by_station else None
                for station_id in station_ids
            ]
            present = [value for value in matrix_row if value is not None]
            if len(present) < 2:
                continue

            formation = str(service.get("numero") or service.get("id") or "").strip()
            if not formation:
                continue
            identity = "|".join([
                formation,
                str(direction["direction"]),
                *(str(minute_value(value)) for value in matrix_row),
            ])
            rows.setdefault(identity, (formation, matrix_row))

        ordered = sorted(
            rows.values(),
            key=lambda pair: next(
                minute_value(value) for value in pair[1] if value is not None
            ),
        )
        snapshot = ScheduleSnapshot(
            stations=[station["name"] for station in direction["stations"]],
            formations=[formation for formation, _ in ordered],
            matrix=[row for _, row in ordered],
            source_date=source_date,
        )
        validate_snapshot(snapshot, bool(direction.get("allow_short_turns")))
        return snapshot

    @staticmethod
    def _response_items(response: Any) -> list[dict[str, Any]]:
        if isinstance(response, list):
            return response
        for key in ("results", "resultado", "data"):
            if isinstance(response, dict) and isinstance(response.get(key), list):
                return response[key]
        return []

    def _discover_direction(
        self,
        branch_id: int,
        direction_number: int,
        origin_id: int,
        destination_id: int,
        source_date: dt.date,
    ) -> dict[str, Any]:
        response = self.get(
            f"/arribos/estacion/{origin_id}",
            {
                "hasta": destination_id,
                "fecha": source_date.isoformat(),
                "hora": "00:00",
                "cantidad": 100,
                "paraApp": "true",
                "ramal": branch_id,
                "sentido": direction_number,
            },
        )
        itineraries: list[list[dict[str, Any]]] = []
        for item in self._response_items(response):
            service = item.get("servicio", item)
            try:
                item_direction = int(service.get("sentido", direction_number))
            except (TypeError, ValueError):
                continue
            stations = [
                station for station in service.get("estaciones", [])
                if station.get("idElemento") is not None and station.get("nombre")
            ]
            if item_direction == direction_number and len(stations) >= 2:
                itineraries.append(stations)
        if not itineraries:
            raise SourceUnavailable(
                f"SOFSE no publicó itinerarios del ramal {branch_id}, sentido "
                f"{direction_number}, para {source_date.isoformat()}"
            )
        stations = max(itineraries, key=len)
        return {
            "file_suffix": compact_suffix(stations[-1]["nombre"]),
            "destination": str(stations[-1]["nombre"]),
            "query_station_id": int(stations[0]["idElemento"]),
            "destination_id": int(stations[-1]["idElemento"]),
            "branch_id": branch_id,
            "direction": direction_number,
            "allow_short_turns": True,
            "stations": [
                {"name": str(station["nombre"]), "id": int(station["idElemento"])}
                for station in stations
            ],
        }

    def discover_route_config(
        self,
        selection: str,
        catalog_row: dict[str, str],
        source_date: dt.date,
    ) -> dict[str, Any]:
        try:
            branch_id = int(re.search(r"\d+", catalog_row["ID API"]).group())
        except (AttributeError, KeyError, ValueError) as exc:
            raise ConfigurationError(
                f"ID SOFSE inválido para {selection}: {catalog_row.get('ID API')}"
            ) from exc

        stations_response = self.get(
            "/infraestructura/estaciones", {"idRamal": branch_id},
        )
        stations = self._response_items(stations_response)
        if len(stations) < 2:
            raise SourceUnavailable(f"SOFSE no devolvió estaciones para {selection}")

        branch_name = catalog_row.get("Nombre API") or selection
        endpoint_names = [
            part.strip() for part in re.split(r"\s*[-–—]\s*", branch_name)
            if part.strip()
        ]

        def station_for(name: str, fallback: dict[str, Any]) -> dict[str, Any]:
            wanted = normalize_text(name)
            ranked = sorted(
                stations,
                key=lambda station: difflib.SequenceMatcher(
                    None, wanted, normalize_text(station.get("nombre")),
                ).ratio(),
                reverse=True,
            )
            return ranked[0] if ranked else fallback

        first = station_for(endpoint_names[0], stations[0]) if endpoint_names else stations[0]
        last = station_for(endpoint_names[-1], stations[-1]) if endpoint_names else stations[-1]
        try:
            first_id = int(first["id_estacion"])
            last_id = int(last["id_estacion"])
        except (KeyError, TypeError, ValueError) as exc:
            raise SourceUnavailable(
                f"SOFSE no devolvió IDs de cabecera utilizables para {selection}"
            ) from exc

        directions = [
            self._discover_direction(branch_id, 1, first_id, last_id, source_date),
            self._discover_direction(branch_id, 2, last_id, first_id, source_date),
        ]
        return {
            "id": f"tren-{slugify(selection)}",
            "provider": "sofse",
            "folder": f"Horarios/Trenes/{folder_component(selection)}",
            "type": "Tren",
            "branch": catalog_row.get("Linea / ramal") or "Trenes Argentinos",
            "company": "Trenes Argentinos",
            "route": selection,
            "website": (
                "https://www.argentina.gob.ar/transporte/trenes-argentinos/"
                "horarios-tarifas-y-recorridos-de-trenes"
            ),
            "source_url": f"{SOFSE_BASE_URL}/arribos/estacion/{first_id}",
            "days": ["Laboral", "Sabado", "Domingo"],
            "directions": directions,
        }


class CuandoSuboProvider:
    def __init__(self, http: JsonHttpClient):
        self.http = http
        self.cache: dict[tuple[str, str], Any] = {}

    def schedule_for_stop(self, stop_id: str, source_date: dt.date) -> Any:
        cache_key = (stop_id, source_date.isoformat())
        if cache_key in self.cache:
            return self.cache[cache_key]
        query = urllib.parse.urlencode({
            "date": source_date.isoformat(),
            "key": CUANDO_SUBO_API_KEY,
        })
        response = self.http.request(
            f"{CUANDO_SUBO_BASE_URL}/schedule-for-stop/{stop_id}.json?{query}"
        )
        if response.get("code") != 200:
            raise SourceUnavailable(
                f"Cuándo SUBO respondio {response.get('code')}: {response.get('text')}"
            )
        self.cache[cache_key] = response
        return response

    @staticmethod
    def route_times(
        response: dict[str, Any], route_id: str, source_date: dt.date,
    ) -> dict[str, int]:
        result: dict[str, int] = {}
        schedules = response.get("data", {}).get("entry", {}).get(
            "stopRouteSchedules", []
        )
        for route_schedule in schedules:
            if route_schedule.get("routeId") != route_id:
                continue
            for direction_schedule in route_schedule.get(
                "stopRouteDirectionSchedules", []
            ):
                for stop_time in direction_schedule.get("scheduleStopTimes", []):
                    timestamp = (
                        stop_time.get("departureTime")
                        if stop_time.get("departureEnabled")
                        else stop_time.get("arrivalTime")
                    )
                    parsed = time_from_epoch_millis(timestamp, source_date)
                    trip_id = str(stop_time.get("tripId") or "").strip()
                    if parsed is not None and trip_id:
                        result[trip_id] = parsed
        return result

    def snapshot(
        self, route: dict[str, Any], direction: dict[str, Any], source_date: dt.date,
    ) -> ScheduleSnapshot:
        station_maps: list[dict[str, int]] = []
        for station in direction["stations"]:
            response = self.schedule_for_stop(station["stop_id"], source_date)
            station_maps.append(self.route_times(response, direction["route_id"], source_date))

        origin_trips = station_maps[0]
        if not origin_trips:
            raise IncompleteSchedule(
                f"Sin servicios {direction['route_id']} en {source_date.isoformat()}"
            )

        rows: list[tuple[str, list[int | None]]] = []
        for trip_id, origin_time in origin_trips.items():
            row = [mapping.get(trip_id) for mapping in station_maps]
            if row[-1] is None:
                continue
            rows.append((trip_id, row))
        rows.sort(key=lambda pair: minute_value(pair[1][0]) or 0)

        snapshot = ScheduleSnapshot(
            stations=[station["name"] for station in direction["stations"]],
            formations=[trip_id for trip_id, _ in rows],
            matrix=[row for _, row in rows],
            source_date=source_date,
        )
        validate_snapshot(snapshot, False)
        return snapshot

    def stops_for_route(self, route_id: str) -> tuple[dict[str, Any], list[dict[str, str]]]:
        query = urllib.parse.urlencode({"key": CUANDO_SUBO_API_KEY})
        response = self.http.request(
            f"{CUANDO_SUBO_BASE_URL}/stops-for-route/{route_id}.json?{query}"
        )
        if response.get("code") != 200:
            raise SourceUnavailable(
                f"Cuándo SUBO no pudo describir {route_id}: {response.get('text')}"
            )
        data = response.get("data", {})
        entry = data.get("entry") or {}
        references = data.get("references") or {}
        stops_by_id = {
            str(stop.get("id")): stop
            for stop in references.get("stops", [])
            if stop.get("id")
        }
        groups = []
        for grouping in entry.get("stopGroupings", []):
            for group in grouping.get("stopGroups", []):
                stop_ids = [str(value) for value in group.get("stopIds", [])]
                if stop_ids:
                    groups.append((stop_ids, group))
        ordered_ids = max(groups, key=lambda pair: len(pair[0]))[0] if groups else [
            str(value) for value in entry.get("stopIds", [])
        ]
        ordered_stops = [
            {
                "name": str(stops_by_id.get(stop_id, {}).get("name") or stop_id),
                "stop_id": stop_id,
            }
            for stop_id in ordered_ids
        ]
        target_route = next(
            (
                route for route in references.get("routes", [])
                if str(route.get("id")) == route_id
            ),
            {"id": route_id},
        )
        return target_route, ordered_stops

    @staticmethod
    def _sample_stops(stops: list[dict[str, str]]) -> list[dict[str, str]]:
        """Limita llamadas conservando cabeceras y puntos uniformes del recorrido."""
        if len(stops) <= MAX_AUTODISCOVERED_BUS_STOPS:
            return stops
        last = len(stops) - 1
        indexes = {
            round(index * last / (MAX_AUTODISCOVERED_BUS_STOPS - 1))
            for index in range(MAX_AUTODISCOVERED_BUS_STOPS)
        }
        return [stops[index] for index in sorted(indexes)]

    def discover_route_config(
        self,
        selection: str,
        catalog_row: dict[str, str],
    ) -> dict[str, Any]:
        route_ids = list(dict.fromkeys(re.findall(r"\d+_\d+", catalog_row.get("ID API", ""))))
        if not route_ids:
            raise ConfigurationError(
                f"La fila de catálogo de {selection} no contiene IDs OneBusAway"
            )
        directions: list[dict[str, Any]] = []
        for route_id in route_ids:
            route_info, all_stops = self.stops_for_route(route_id)
            if len(all_stops) < 2:
                raise SourceUnavailable(f"{route_id} no contiene suficientes paradas")
            description = str(route_info.get("description") or "")
            destination = (
                description.rsplit(":", 1)[-1].strip()
                if ":" in description else all_stops[-1]["name"]
            )
            directions.append({
                "file_suffix": compact_suffix(destination),
                "destination": destination,
                "route_id": route_id,
                "stations": self._sample_stops(all_stops),
                "discovered_total_stops": len(all_stops),
            })
        suffixes: set[str] = set()
        for direction in directions:
            suffix = direction["file_suffix"]
            if suffix in suffixes:
                direction["file_suffix"] = f"{suffix}{direction['route_id'].split('_')[-1]}"
            suffixes.add(direction["file_suffix"])
        return {
            "id": f"colectivo-{slugify(selection)}",
            "provider": "cuando_subo",
            "folder": f"Horarios/Colectivos/{folder_component(selection)}",
            "type": "Colectivo",
            "branch": catalog_row.get("Linea / ramal") or selection,
            "company": catalog_row.get("Empresa / agencia") or "Operador publicado por SUBE",
            "route": selection,
            "website": "https://www.argentina.gob.ar/sube/cuandosubo",
            "source_url": CUANDO_SUBO_BASE_URL,
            "days": ["Laboral", "Sabado", "Domingo"],
            "directions": directions,
        }


def safe_table_name(route_id: str, day: str, suffix: str) -> str:
    raw = normalize_text(f"{route_id}_{day}_{suffix}")
    cleaned = re.sub(r"[^a-z0-9_]", "_", raw).strip("_")
    return f"T_{cleaned}"[:250]


def create_workbook(
    path: Path,
    route: dict[str, Any],
    direction: dict[str, Any],
    day: str,
    snapshot: ScheduleSnapshot,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Horarios"

    metadata = [
        (1, "Tipo", route["type"]),
        (3, "Ramal", route["branch"]),
        (5, "Empresa", route["company"]),
        (7, "Recorrido", route["route"]),
        (9, "Dia", day),
        (11, "Sentido", direction["destination"]),
        (13, "Estaciones", len(snapshot.stations)),
        (15, "Formaciones", len(snapshot.formations)),
        (17, "Vigencia", snapshot.source_date),
        (19, "Website", route["website"]),
        (21, "Link", route["source_url"]),
        (23, "Metodo", "API"),
    ]
    for row, label, value in metadata:
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row + 1, column=1, value=value)

    sheet.cell(row=1, column=2, value="Formacion")
    for index, station in enumerate(snapshot.stations, start=3):
        sheet.cell(row=1, column=index, value=station)

    for row_index, (formation, times) in enumerate(
        zip(snapshot.formations, snapshot.matrix), start=2,
    ):
        sheet.cell(row=row_index, column=2, value=formation)
        for column_index, value in enumerate(times, start=3):
            if value is None:
                cell = sheet.cell(row=row_index, column=column_index, value="null")
            elif value < 1440:
                cell = sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=dt.time(value // 60, value % 60),
                )
                cell.number_format = "h:mm"
            else:
                cell = sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=f"{value // 60:02d}:{value % 60:02d}",
                )

    used_rows = max(24, len(snapshot.formations) + 1)
    used_columns = len(snapshot.stations) + 2
    body = sheet.cell(1, 1).offset(
        row=used_rows - 1, column=used_columns - 1,
    ).coordinate
    table = Table(
        displayName=safe_table_name(route["id"], day, direction["file_suffix"]),
        ref=f"A1:{body}",
    )
    sheet.add_table(table)

    for row in sheet.iter_rows(
        min_row=1, max_row=used_rows, min_col=1, max_col=used_columns,
    ):
        for cell in row:
            cell.font = Font(name="Aptos Narrow", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell_ref in ("A20", "A22"):
        cell = sheet[cell_ref]
        cell.hyperlink = str(cell.value)
        cell.style = "Hyperlink"
        cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet["A18"].number_format = "yyyy-mm-dd"
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 20
    for column in range(3, used_columns + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 21
    sheet.freeze_panes = "C2"

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def read_existing_method(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return normalize_text(workbook.active["A24"].value)
    finally:
        workbook.close()


def validate_generated_workbook(path: Path) -> dict[str, Any]:
    try:
        from procesar_horarios import parse_file
    except ImportError:
        import sys
        sys.path.insert(0, str(PROJECT_ROOT))
        from procesar_horarios import parse_file
    return parse_file(path)


def update_target(
    provider: Any,
    route: dict[str, Any],
    direction: dict[str, Any],
    day: str,
    weekday: int,
    today: dt.date,
    *,
    dry_run: bool,
) -> tuple[str, str, Path]:
    source_date = next_weekday(today, weekday)
    output = PROJECT_ROOT / route["folder"] / f"{day}{direction['file_suffix']}.xlsx"
    snapshot = provider.snapshot(route, direction, source_date)

    output.parent.mkdir(parents=True, exist_ok=True)
    candidate: Path | None = None
    try:
        # El archivo temporal debe nacer directamente en la carpeta destino.
        # En Windows, moverlo desde un subdirectorio temporal puede conservar
        # una ACL restrictiva y hacer que Excel informe una falsa extensión inválida.
        with tempfile.NamedTemporaryFile(
            prefix=".solaris-xlsx-",
            suffix=".xlsx",
            dir=output.parent,
            delete=False,
        ) as temp_file:
            candidate = Path(temp_file.name)
        create_workbook(candidate, route, direction, day, snapshot)
        parsed_candidate = validate_generated_workbook(candidate)

        if output.exists():
            parsed_existing = validate_generated_workbook(output)
            if (
                parsed_existing["hash"] == parsed_candidate["hash"]
                and read_existing_method(output) == "api"
            ):
                return "unchanged", "Cronograma y metodo sin cambios", output

        if dry_run:
            return "would_update", f"Snapshot valido de {source_date}", output

        os.replace(candidate, output)
        candidate = None
    finally:
        if candidate is not None:
            candidate.unlink(missing_ok=True)
    return "updated", f"Actualizado desde API con fecha {source_date}", output


def load_config(path: Path) -> dict[str, Any]:
    try:
        config = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ConfigurationError(f"No se pudo leer {path}: {exc}") from exc
    if config.get("version") != 1:
        raise ConfigurationError("Version de configuracion no soportada")
    if not config.get("routes") or not config.get("days"):
        raise ConfigurationError("La configuracion no contiene rutas o dias")
    return config


def resolve_requested_routes(
    config: dict[str, Any],
    branches_path: Path,
    providers: dict[str, Any],
    today: dt.date,
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    selections = load_branch_selections(branches_path)
    catalog = load_branch_catalog(branches_path)
    configured = config["routes"]
    resolved: list[dict[str, Any]] = []
    unresolved: list[dict[str, str]] = []
    seen: set[str] = set()
    discovery_date = next_weekday(today, int(config["days"].get("Laboral", 0)))

    for type_name, values in selections.items():
        for selection in values:
            wanted = normalize_text(selection)
            route = next(
                (
                    candidate for candidate in configured
                    if normalize_text(candidate.get("type")) == normalize_text(type_name)
                    and wanted in _configured_route_names(candidate)
                ),
                None,
            )
            try:
                if route is None:
                    catalog_row = _catalog_match(selection, type_name, catalog)
                    if catalog_row is None:
                        raise SourceUnavailable(
                            "No hubo una coincidencia única en Lista de ramales"
                        )
                    provider_name = normalize_text(catalog_row.get("Proveedor"))
                    if type_name == "Tren" and provider_name == "sofse":
                        route = providers["sofse"].discover_route_config(
                            selection, catalog_row, discovery_date,
                        )
                    elif type_name == "Colectivo" and provider_name in {
                        "cuando subo", "cuando_subo", "onebusaway",
                    }:
                        route = providers["cuando_subo"].discover_route_config(
                            selection, catalog_row,
                        )
                    else:
                        raise SourceUnavailable(
                            f"Proveedor no soportado en catálogo: {catalog_row.get('Proveedor')}"
                        )
            except (SourceUnavailable, ConfigurationError, KeyError) as exc:
                unresolved.append({
                    "type": type_name,
                    "selection": selection,
                    "message": str(exc),
                })
                logger.warning("RAMAL PENDIENTE %s (%s): %s", selection, type_name, exc)
                continue

            route_key = normalize_text(route["id"])
            if route_key not in seen:
                resolved.append(route)
                seen.add(route_key)
    return resolved, unresolved


def run(
    config_path: Path = DEFAULT_CONFIG,
    *,
    branches_path: Path = DEFAULT_BRANCHES,
    today: dt.date | None = None,
    dry_run: bool = False,
    only_routes: set[str] | None = None,
) -> dict[str, Any]:
    config = load_config(config_path)
    today = today or dt.datetime.now(ARGENTINA_TZ).date()
    http = JsonHttpClient()
    providers = {
        "sofse": SofseProvider(http),
        "cuando_subo": CuandoSuboProvider(http),
    }
    routes, unresolved = resolve_requested_routes(
        config, branches_path, providers, today,
    )
    summary: dict[str, Any] = {
        "updated": 0,
        "unchanged": 0,
        "would_update": 0,
        "preserved": 0,
        "unresolved": len(unresolved),
        "targets": [],
        "unresolved_routes": unresolved,
    }

    for route in routes:
        if only_routes and route["id"] not in only_routes:
            continue
        if route.get("provider") not in providers:
            raise ConfigurationError(
                f"Proveedor desconocido en {route.get('id')}: {route.get('provider')}"
            )
        provider = providers[route["provider"]]
        for day in route["days"]:
            if day not in config["days"]:
                raise ConfigurationError(f"Dia sin mapeo: {day}")
            for direction in route["directions"]:
                output = (
                    PROJECT_ROOT / route["folder"]
                    / f"{day}{direction['file_suffix']}.xlsx"
                )
                try:
                    status, message, output = update_target(
                        provider,
                        route,
                        direction,
                        day,
                        int(config["days"][day]),
                        today,
                        dry_run=dry_run,
                    )
                except SourceUnavailable as exc:
                    status = "preserved"
                    message = str(exc)
                    logger.warning("PRESERVADO %s: %s", output, message)
                else:
                    logger.info("%s %s: %s", status.upper(), output, message)

                summary[status] += 1
                summary["targets"].append({
                    "route": route["id"],
                    "day": day,
                    "direction": direction["destination"],
                    "file": str(output.relative_to(PROJECT_ROOT)),
                    "status": status,
                    "message": message,
                })

    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--branches", type=Path, default=DEFAULT_BRANCHES)
    parser.add_argument("--today", type=dt.date.fromisoformat)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--route", action="append", dest="routes")
    parser.add_argument("--verbose", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    try:
        summary = run(
            args.config,
            branches_path=args.branches,
            today=args.today,
            dry_run=args.dry_run,
            only_routes=set(args.routes) if args.routes else None,
        )
    except ConfigurationError as exc:
        logger.error("Configuracion invalida: %s", exc)
        return 2
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
