#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
SolarisPKN-Transport V2
Importador XLSX -> SQLite

Arquitectura congelada:
- Recorrido + Día + Sentido = GRILLA
- Estado vigente, sin historial
- Fuente oficial pertenece al recorrido
- Ausencia de horario = ausencia de fila
- Cruces de medianoche = minutos mayores o iguales a 1440
- Cada grilla conserva si su snapshot fue Manual o API
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import logging
import math
import os
import re
import sqlite3
import tempfile
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("solaris")

DB_PATH = Path("horarios.db")
HORARIOS_DIR = Path("Horarios")

LABEL_ALIASES = {
    "tipo": "tipo",
    "ramal": "ramal",
    "empresa": "empresa",
    "recorrido": "recorrido",
    "dia": "dia",
    "sentido": "sentido",
    "estaciones": "estaciones",
    "formaciones": "formaciones",
    "vigencia": "vigencia",
    "website": "website",
    "link": "link",
    "metodo": "metodo",
}

REQUIRED_METADATA = (
    "tipo", "ramal", "empresa", "recorrido", "dia", "sentido",
    "estaciones", "formaciones", "vigencia", "website", "link", "metodo",
)

DAY_MAP = {
    "laboral": "Lunes a Viernes",
    "nolaboral": "No Laboral",
    "sabado": "Sábado",
    "domingo": "Domingo",
    "feriados": "Feriados",
}

ABSENCE_STRINGS = {"", "null", "none", "n/a", "-", "—", "nan"}

DDL = """
CREATE TABLE IF NOT EXISTS recorridos (
    id INTEGER PRIMARY KEY,
    nombre TEXT NOT NULL,
    nombre_norm TEXT NOT NULL,
    tipo TEXT NOT NULL,
    tipo_norm TEXT NOT NULL,
    ramal TEXT NOT NULL,
    ramal_norm TEXT NOT NULL,
    empresa TEXT NOT NULL,
    empresa_norm TEXT NOT NULL,
    website_url TEXT NOT NULL,
    pdf_url TEXT NOT NULL,
    vigencia_texto TEXT NOT NULL,
    vigencia_iso TEXT NOT NULL,
    creado_en TEXT DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(tipo_norm, ramal_norm, empresa_norm, nombre_norm)
);

CREATE TABLE IF NOT EXISTS estaciones (
    id INTEGER PRIMARY KEY,
    nombre TEXT NOT NULL UNIQUE,
    nombre_norm TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS dias (
    id INTEGER PRIMARY KEY,
    nombre TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS grillas (
    id INTEGER PRIMARY KEY,
    recorrido_id INTEGER NOT NULL REFERENCES recorridos(id),
    dia_id INTEGER NOT NULL REFERENCES dias(id),
    sentido_estacion_id INTEGER NOT NULL REFERENCES estaciones(id),
    hash_contenido TEXT NOT NULL,
    metodo_actualizacion TEXT NOT NULL DEFAULT 'Manual',
    fuente_archivo TEXT,
    actualizado_en TEXT DEFAULT CURRENT_TIMESTAMP,
    creado_en TEXT DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(recorrido_id, dia_id, sentido_estacion_id)
);

CREATE TABLE IF NOT EXISTS grilla_estaciones (
    id INTEGER PRIMARY KEY,
    grilla_id INTEGER NOT NULL REFERENCES grillas(id) ON DELETE CASCADE,
    estacion_id INTEGER NOT NULL REFERENCES estaciones(id),
    orden INTEGER NOT NULL,
    UNIQUE(grilla_id, estacion_id),
    UNIQUE(grilla_id, orden)
);

CREATE TABLE IF NOT EXISTS grilla_formaciones (
    id INTEGER PRIMARY KEY,
    grilla_id INTEGER NOT NULL REFERENCES grillas(id) ON DELETE CASCADE,
    nombre TEXT NOT NULL,
    nombre_norm TEXT NOT NULL,
    orden INTEGER NOT NULL,
    UNIQUE(grilla_id, nombre_norm),
    UNIQUE(grilla_id, orden)
);

CREATE TABLE IF NOT EXISTS horarios (
    id INTEGER PRIMARY KEY,
    grilla_formacion_id INTEGER NOT NULL REFERENCES grilla_formaciones(id) ON DELETE CASCADE,
    grilla_estacion_id INTEGER NOT NULL REFERENCES grilla_estaciones(id) ON DELETE CASCADE,
    minutos INTEGER NOT NULL,
    UNIQUE(grilla_formacion_id, grilla_estacion_id)
);

CREATE TABLE IF NOT EXISTS import_logs (
    id INTEGER PRIMARY KEY,
    archivo TEXT NOT NULL,
    estado TEXT NOT NULL,
    mensaje TEXT,
    creado_en TEXT DEFAULT CURRENT_TIMESTAMP
);
"""


class ScheduleImportError(Exception):
    pass


def new_stats() -> dict:
    return {
        "found": 0, "ok": 0, "skip": 0, "rejected": 0,
        "grids_created": 0, "grids_updated": 0,
        "timetables_inserted": 0, "errors": 0,
    }


# ------------------------------------------------------------
# Normalización
# ------------------------------------------------------------

def normalize_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    s = str(value)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower()
    out = []
    for ch in s:
        cat = unicodedata.category(ch)
        if cat.startswith("P") or cat.startswith("S") or cat.startswith("C") or cat.startswith("Z"):
            out.append(" ")
        else:
            out.append(ch)
    s = "".join(out)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_day(value) -> str:
    key = normalize_text(value)
    key_no_spaces = key.replace(" ", "")
    if key in DAY_MAP:
        return DAY_MAP[key]
    if key_no_spaces in DAY_MAP:
        return DAY_MAP[key_no_spaces]
    raise ScheduleImportError(f"Día no soportado: {value!r}")


# ------------------------------------------------------------
# Parseo de valores simples
# ------------------------------------------------------------

def is_absence(value) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str):
        return value.strip().lower() in ABSENCE_STRINGS
    return False


def parse_positive_int(value, field_name: str) -> int:
    if isinstance(value, bool):
        raise ScheduleImportError(f"{field_name} inválido: booleano")
    if isinstance(value, int):
        v = value
    elif isinstance(value, float):
        if not math.isfinite(value) or not value.is_integer():
            raise ScheduleImportError(f"{field_name} inválido: {value!r}")
        v = int(value)
    elif isinstance(value, str):
        s = value.strip()
        if not s:
            raise ScheduleImportError(f"{field_name} vacío")
        try:
            v = int(s)
        except ValueError:
            raise ScheduleImportError(f"{field_name} inválido: {value!r}")
    else:
        raise ScheduleImportError(f"{field_name} inválido: {value!r}")
    if v <= 0:
        raise ScheduleImportError(f"{field_name} debe ser positivo")
    return v


def parse_formation(value) -> tuple[str, str]:
    if value is None:
        raise ScheduleImportError("Formación vacía")
    if isinstance(value, bool):
        raise ScheduleImportError("Formación inválida: booleano")
    if isinstance(value, int):
        s = str(value)
    elif isinstance(value, float):
        if not math.isfinite(value):
            raise ScheduleImportError("Formación inválida: float no finito")
        if value.is_integer():
            s = str(int(value))
        else:
            raise ScheduleImportError(f"Formación inválida: {value!r}")
    elif isinstance(value, str):
        s = value.strip()
        if not s:
            raise ScheduleImportError("Formación vacía")
    else:
        s = str(value).strip()
        if not s:
            raise ScheduleImportError("Formación vacía")
    norm = normalize_text(s)
    if not norm:
        raise ScheduleImportError(f"Formación inválida: {value!r}")
    return s, norm


def parse_time_to_minutes(value) -> int:
    if isinstance(value, bool):
        raise ScheduleImportError("Horario inválido: booleano")
    if isinstance(value, dt.time):
        return value.hour * 60 + value.minute
    if isinstance(value, dt.datetime):
        return value.hour * 60 + value.minute
    if isinstance(value, dt.timedelta):
        total_seconds = value.total_seconds()
        if total_seconds < 0 or total_seconds >= 86400:
            raise ScheduleImportError("Horario fuera de rango")
        return int(total_seconds // 60)
    if isinstance(value, int):
        if value == 0:
            return 0
        raise ScheduleImportError(f"Horario inválido: {value!r}")
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ScheduleImportError("Horario inválido: float no finito")
        frac = value - int(value)
        if frac < 0:
            frac += 1
        minutes = int(math.floor(frac * 1440 + 1e-7))
        if minutes < 0 or minutes >= 1440:
            raise ScheduleImportError("Horario fuera de rango")
        return minutes
    if isinstance(value, str):
        s = value.strip()
        m = re.fullmatch(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", s)
        if not m:
            raise ScheduleImportError(f"Horario inválido: {value!r}")
        hour = int(m.group(1))
        minute = int(m.group(2))
        second = int(m.group(3) or 0)
        if hour > 47 or minute > 59 or second > 59:
            raise ScheduleImportError(f"Horario inválido: {value!r}")
        return hour * 60 + minute
    raise ScheduleImportError(f"Horario inválido: {value!r}")


def parse_vigencia_iso(value) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if not isinstance(value, str):
        raise ScheduleImportError(f"Vigencia inválida: {value!r}")

    s = value.strip()
    if not s:
        raise ScheduleImportError("Vigencia vacía")

    # ISO: YYYY-MM-DD o YYYY/MM/DD, opcionalmente con hora
    # Nota: (?:...) es non-capturing group, así que m.groups() devuelve exactamente 3 elementos.
    # Se usa m.group(N) explícitamente para máxima claridad.
    m = re.fullmatch(
        r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})(?:[ T]\d{1,2}:\d{2}(?::\d{2})?)?",
        s,
    )
    if m:
        y = int(m.group(1))
        mo = int(m.group(2))
        d = int(m.group(3))
        try:
            return dt.date(y, mo, d).isoformat()
        except ValueError:
            raise ScheduleImportError(f"Vigencia inválida: {value!r}")

    # Formato corto con barras: D/M/YY o M/D/YY
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", s)
    if not m:
        raise ScheduleImportError(f"Vigencia inválida: {value!r}")

    a, b, y_raw = m.groups()
    a = int(a)
    b = int(b)
    y_raw_int = int(y_raw)

    if len(y_raw) == 2:
        year = 2000 + y_raw_int if y_raw_int <= 68 else 1900 + y_raw_int
    else:
        year = y_raw_int

    candidates = set()
    for month, day in ((a, b), (b, a)):
        try:
            candidates.add(dt.date(year, month, day))
        except ValueError:
            pass

    if len(candidates) == 1:
        return candidates.pop().isoformat()
    if len(candidates) > 1:
        raise ScheduleImportError(f"Vigencia ambigua: {value!r}")
    raise ScheduleImportError(f"Vigencia inválida: {value!r}")


# ------------------------------------------------------------
# Lectura del XLSX
# ------------------------------------------------------------

def open_workbook(path: Path):
    return load_workbook(path, data_only=True, read_only=False)


def find_matrix_header_rows(ws) -> list[int]:
    rows = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=2).value
        if normalize_text(val) == "formacion":
            rows.append(r)
    return rows


def find_schedule_sheet(wb):
    candidates = []
    for ws in wb.worksheets:
        rows = find_matrix_header_rows(ws)
        if len(rows) > 1:
            raise ScheduleImportError("Más de una matriz detectada en la misma hoja")
        if len(rows) == 1:
            candidates.append((ws, rows[0]))
    if len(candidates) == 0:
        raise ScheduleImportError("Ninguna hoja contiene una matriz válida")
    if len(candidates) > 1:
        raise ScheduleImportError("Más de una hoja contiene matriz válida")
    return candidates[0]


def extract_metadata(ws) -> dict:
    metadata = {}
    consumed_rows = set()
    for r in range(1, ws.max_row + 1):
        if r in consumed_rows:
            continue
        raw = ws.cell(row=r, column=1).value
        norm = normalize_text(raw)
        if norm not in LABEL_ALIASES:
            continue
        label = LABEL_ALIASES[norm]
        if label in metadata:
            raise ScheduleImportError(f"Metadata duplicada: {label}")
        value_row = r + 1
        if value_row > ws.max_row:
            raise ScheduleImportError(f"Metadata incompleta: {label}")
        value = ws.cell(row=value_row, column=1).value
        value_norm = normalize_text(value)
        if value_norm == "":
            raise ScheduleImportError(f"Metadata vacía: {label}")
        if value_norm in LABEL_ALIASES:
            raise ScheduleImportError(f"Metadata inválida: {label} no tiene valor")
        metadata[label] = value
        consumed_rows.add(value_row)
    return metadata


def parse_stations(ws, header_row: int, m: int) -> list[dict]:
    stations = []
    seen_norms = set()
    for col in range(3, 3 + m):
        raw = ws.cell(row=header_row, column=col).value
        if raw is None:
            raise ScheduleImportError(f"Estación vacía en columna {col}")
        original = str(raw).strip()
        norm = normalize_text(raw)
        if not norm:
            raise ScheduleImportError(f"Estación inválida en columna {col}")
        if norm in seen_norms:
            raise ScheduleImportError(f"Estación duplicada tras normalización: {norm}")
        seen_norms.add(norm)
        stations.append({"original": original, "norm": norm})
    for col in range(3 + m, ws.max_column + 1):
        raw = ws.cell(row=header_row, column=col).value
        if raw is not None and str(raw).strip() != "":
            raise ScheduleImportError(
                f"M={m} pero se detectó contenido adicional de estaciones en columna {col}"
            )
    return stations


def parse_formations(ws, header_row: int, n: int) -> list[dict]:
    formations = []
    seen_norms = set()
    for i in range(n):
        r = header_row + 1 + i
        raw = ws.cell(row=r, column=2).value
        original, norm = parse_formation(raw)
        if norm in seen_norms:
            raise ScheduleImportError(f"Formación duplicada tras normalización: {norm}")
        seen_norms.add(norm)
        formations.append({"original": original, "norm": norm, "row": r})
    for r in range(header_row + 1 + n, ws.max_row + 1):
        raw = ws.cell(row=r, column=2).value
        if raw is not None and str(raw).strip() != "":
            raise ScheduleImportError(
                f"N={n} pero se detectó contenido adicional de formaciones en fila {r}"
            )
    return formations


def parse_matrix(ws, header_row: int, m: int, formations: list[dict]) -> list[list[int | None]]:
    for f in formations:
        for col in range(3 + m, ws.max_column + 1):
            raw = ws.cell(row=f["row"], column=col).value
            if raw is not None and str(raw).strip() != "":
                raise ScheduleImportError(
                    f"Contenido adicional fuera de la matriz en fila {f['row']} columna {col}"
                )
    matrix = []
    for f in formations:
        row_values = []
        last_present = None
        for j in range(m):
            col = 3 + j
            raw = ws.cell(row=f["row"], column=col).value
            if is_absence(raw):
                row_values.append(None)
                continue
            minutes = parse_time_to_minutes(raw)
            if last_present is not None and minutes < last_present:
                # Excel representa 00:xx igual en cualquier fecha. Unicamente
                # inferimos cambio de dia en la ventana nocturna para no ocultar
                # un verdadero error de orden en horarios diurnos.
                if last_present >= 18 * 60 and minutes <= 6 * 60:
                    minutes += 1440
                else:
                    raise ScheduleImportError(
                        "Horario decreciente dentro de la secuencia de estaciones"
                    )
            last_present = minutes
            row_values.append(minutes)
        matrix.append(row_values)
    return matrix


def validate_grid(stations, formations, matrix, sentido_norm: str) -> None:
    if len(matrix) != len(formations):
        raise ScheduleImportError("La matriz no coincide con la cantidad de formaciones")
    for row in matrix:
        if len(row) != len(stations):
            raise ScheduleImportError("La matriz no coincide con la cantidad de estaciones")
    if sentido_norm not in {s["norm"] for s in stations}:
        raise ScheduleImportError("El sentido no corresponde a una estación de la matriz")


def build_grid_payload(recorrido, dia_canonical, sentido_norm, stations, formations, matrix) -> str:
    lines = [
        "SOLARIS:GRILLA:v1",
        f"R:{normalize_text(recorrido)}",
        f"D:{normalize_text(dia_canonical)}",
        f"S:{sentido_norm}",
        "E:" + "|".join(s["norm"] for s in stations),
        "F:" + "|".join(f["norm"] for f in formations),
        "M:" + ";".join(
            ",".join("-" if cell is None else str(cell) for cell in row)
            for row in matrix
        ),
    ]
    return "\n".join(lines)


def calculate_grid_hash(payload: str) -> str:
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def parse_file(path: Path) -> dict:
    wb = open_workbook(path)
    try:
        ws, header_row = find_schedule_sheet(wb)
        meta = extract_metadata(ws)

        for key in REQUIRED_METADATA:
            if key not in meta:
                raise ScheduleImportError(f"Metadata obligatoria faltante: {key}")
            value = meta[key]
            if value is None or normalize_text(str(value)) == "":
                raise ScheduleImportError(f"Metadata obligatoria vacía: {key}")

        m = parse_positive_int(meta["estaciones"], "Estaciones")
        n = parse_positive_int(meta["formaciones"], "Formaciones")

        stations = parse_stations(ws, header_row, m)

        sentido_original = str(meta["sentido"]).strip()
        sentido_norm = normalize_text(sentido_original)

        formations = parse_formations(ws, header_row, n)
        matrix = parse_matrix(ws, header_row, m, formations)

        dia_canonical = normalize_day(meta["dia"])

        validate_grid(stations, formations, matrix, sentido_norm)

        vigencia_texto = str(meta["vigencia"]).strip()
        vigencia_iso = parse_vigencia_iso(meta["vigencia"])

        website_url = str(meta["website"]).strip()
        pdf_url = str(meta["link"]).strip()
        metodo_actualizacion = str(meta["metodo"]).strip()

        if not website_url:
            raise ScheduleImportError("Website vacío")
        if not pdf_url:
            raise ScheduleImportError("Link vacío")
        if normalize_text(metodo_actualizacion) not in {"manual", "api"}:
            raise ScheduleImportError(
                f"Método de actualización inválido: {metodo_actualizacion!r}"
            )

        tipo = str(meta["tipo"]).strip()
        ramal = str(meta["ramal"]).strip()
        empresa = str(meta["empresa"]).strip()
        recorrido = str(meta["recorrido"]).strip()

        if not tipo:
            raise ScheduleImportError("Tipo vacío")
        if not ramal:
            raise ScheduleImportError("Ramal vacío")
        if not empresa:
            raise ScheduleImportError("Empresa vacía")
        if not recorrido:
            raise ScheduleImportError("Recorrido vacío")

        payload = build_grid_payload(
            recorrido=recorrido,
            dia_canonical=dia_canonical,
            sentido_norm=sentido_norm,
            stations=stations,
            formations=formations,
            matrix=matrix,
        )
        grid_hash = calculate_grid_hash(payload)

        return {
            "tipo": tipo,
            "ramal": ramal,
            "empresa": empresa,
            "recorrido": recorrido,
            "dia_canonical": dia_canonical,
            "sentido_original": sentido_original,
            "sentido_norm": sentido_norm,
            "vigencia_texto": vigencia_texto,
            "vigencia_iso": vigencia_iso,
            "website_url": website_url,
            "pdf_url": pdf_url,
            "metodo_actualizacion": metodo_actualizacion,
            "m": m,
            "n": n,
            "stations": stations,
            "formations": formations,
            "matrix": matrix,
            "payload": payload,
            "hash": grid_hash,
        }
    finally:
        wb.close()


# ------------------------------------------------------------
# Base de datos
# ------------------------------------------------------------

def init_database(db_path=DB_PATH):
    conn = sqlite3.connect(db_path)
    conn.isolation_level = None
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.executescript(DDL)
    ensure_schema_migrations(conn)
    return conn


def ensure_schema_migrations(conn) -> None:
    columns = {
        row["name"] for row in conn.execute("PRAGMA table_info(grillas)").fetchall()
    }
    migrations = {
        "metodo_actualizacion": (
            "ALTER TABLE grillas ADD COLUMN metodo_actualizacion "
            "TEXT NOT NULL DEFAULT 'Manual'"
        ),
        "fuente_archivo": "ALTER TABLE grillas ADD COLUMN fuente_archivo TEXT",
        "actualizado_en": "ALTER TABLE grillas ADD COLUMN actualizado_en TEXT",
    }
    for column, statement in migrations.items():
        if column not in columns:
            conn.execute(statement)


def log_import(conn, archivo: str, estado: str, mensaje: str | None) -> None:
    conn.execute(
        "INSERT INTO import_logs (archivo, estado, mensaje) VALUES (?, ?, ?)",
        (str(archivo), estado, mensaje),
    )
    conn.commit()


def get_or_create_recorrido(
    conn, nombre, nombre_norm, tipo, ramal, empresa,
    website_url, pdf_url, vigencia_texto, vigencia_iso,
) -> tuple[int, sqlite3.Row | None, bool]:
    tipo_norm = normalize_text(tipo)
    ramal_norm = normalize_text(ramal)
    empresa_norm = normalize_text(empresa)

    row = conn.execute(
        """
        SELECT * FROM recorridos
        WHERE tipo_norm = ? AND ramal_norm = ? AND empresa_norm = ? AND nombre_norm = ?
        """,
        (tipo_norm, ramal_norm, empresa_norm, nombre_norm),
    ).fetchone()

    if row:
        return row["id"], row, False

    cur = conn.execute(
        """
        INSERT INTO recorridos (
            nombre, nombre_norm, tipo, tipo_norm, ramal, ramal_norm,
            empresa, empresa_norm, website_url, pdf_url,
            vigencia_texto, vigencia_iso
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (nombre, nombre_norm, tipo, tipo_norm, ramal, ramal_norm,
         empresa, empresa_norm, website_url, pdf_url,
         vigencia_texto, vigencia_iso),
    )
    return cur.lastrowid, None, True


def update_recorrido_if_changed(
    conn, recorrido_id, current_row, nombre, nombre_norm, tipo, ramal, empresa,
    website_url, pdf_url, vigencia_texto, vigencia_iso,
) -> bool:
    if current_row is None:
        return True

    tipo_norm = normalize_text(tipo)
    ramal_norm = normalize_text(ramal)
    empresa_norm = normalize_text(empresa)

    changed = (
        current_row["nombre"] != nombre
        or current_row["nombre_norm"] != nombre_norm
        or current_row["tipo"] != tipo
        or current_row["tipo_norm"] != tipo_norm
        or current_row["ramal"] != ramal
        or current_row["ramal_norm"] != ramal_norm
        or current_row["empresa"] != empresa
        or current_row["empresa_norm"] != empresa_norm
        or current_row["website_url"] != website_url
        or current_row["pdf_url"] != pdf_url
        or current_row["vigencia_texto"] != vigencia_texto
        or current_row["vigencia_iso"] != vigencia_iso
    )

    if changed:
        conn.execute(
            """
            UPDATE recorridos
            SET nombre = ?, nombre_norm = ?, tipo = ?, tipo_norm = ?,
                ramal = ?, ramal_norm = ?, empresa = ?, empresa_norm = ?,
                website_url = ?, pdf_url = ?, vigencia_texto = ?, vigencia_iso = ?
            WHERE id = ?
            """,
            (nombre, nombre_norm, tipo, tipo_norm, ramal, ramal_norm,
             empresa, empresa_norm, website_url, pdf_url,
             vigencia_texto, vigencia_iso, recorrido_id),
        )

    return changed


def get_or_create_dia(conn, dia_canonical: str) -> int:
    row = conn.execute("SELECT id FROM dias WHERE nombre = ?", (dia_canonical,)).fetchone()
    if row:
        return row["id"]
    cur = conn.execute("INSERT INTO dias (nombre) VALUES (?)", (dia_canonical,))
    return cur.lastrowid


def get_or_create_estacion(conn, nombre: str, nombre_norm: str) -> int:
    row = conn.execute(
        "SELECT id FROM estaciones WHERE nombre_norm = ?", (nombre_norm,),
    ).fetchone()
    if row:
        return row["id"]
    cur = conn.execute(
        "INSERT INTO estaciones (nombre, nombre_norm) VALUES (?, ?)",
        (nombre, nombre_norm),
    )
    return cur.lastrowid


def get_or_create_entities(conn, parsed: dict) -> dict:
    nombre_norm = normalize_text(parsed["recorrido"])

    recorrido_id, current_recorrido, created_recorrido = get_or_create_recorrido(
        conn,
        parsed["recorrido"], nombre_norm,
        parsed["tipo"], parsed["ramal"], parsed["empresa"],
        parsed["website_url"], parsed["pdf_url"],
        parsed["vigencia_texto"], parsed["vigencia_iso"],
    )

    if created_recorrido:
        source_changed = True
    else:
        source_changed = update_recorrido_if_changed(
            conn, recorrido_id, current_recorrido,
            parsed["recorrido"], nombre_norm,
            parsed["tipo"], parsed["ramal"], parsed["empresa"],
            parsed["website_url"], parsed["pdf_url"],
            parsed["vigencia_texto"], parsed["vigencia_iso"],
        )

    dia_id = get_or_create_dia(conn, parsed["dia_canonical"])

    station_ids = []
    for st in parsed["stations"]:
        station_ids.append(get_or_create_estacion(conn, st["original"], st["norm"]))

    sentido_index = next(
        i for i, st in enumerate(parsed["stations"]) if st["norm"] == parsed["sentido_norm"]
    )
    sentido_estacion_id = station_ids[sentido_index]

    return {
        "recorrido_id": recorrido_id,
        "source_changed": source_changed,
        "dia_id": dia_id,
        "station_ids": station_ids,
        "sentido_estacion_id": sentido_estacion_id,
    }


def delete_grid_content(conn, grilla_id: int) -> None:
    conn.execute(
        """
        DELETE FROM horarios
        WHERE grilla_formacion_id IN (
            SELECT id FROM grilla_formaciones WHERE grilla_id = ?
        )
        OR grilla_estacion_id IN (
            SELECT id FROM grilla_estaciones WHERE grilla_id = ?
        )
        """,
        (grilla_id, grilla_id),
    )
    conn.execute("DELETE FROM grilla_formaciones WHERE grilla_id = ?", (grilla_id,))
    conn.execute("DELETE FROM grilla_estaciones WHERE grilla_id = ?", (grilla_id,))


def load_grid(conn, grilla_id: int, parsed: dict, station_ids: list[int]) -> int:
    grilla_estacion_ids = []
    for orden, estacion_id in enumerate(station_ids, start=1):
        cur = conn.execute(
            "INSERT INTO grilla_estaciones (grilla_id, estacion_id, orden) VALUES (?, ?, ?)",
            (grilla_id, estacion_id, orden),
        )
        grilla_estacion_ids.append(cur.lastrowid)

    grilla_formacion_ids = []
    for orden, f in enumerate(parsed["formations"], start=1):
        cur = conn.execute(
            "INSERT INTO grilla_formaciones (grilla_id, nombre, nombre_norm, orden) VALUES (?, ?, ?, ?)",
            (grilla_id, f["original"], f["norm"], orden),
        )
        grilla_formacion_ids.append(cur.lastrowid)

    horarios_insertados = 0
    for f_idx, row in enumerate(parsed["matrix"]):
        formacion_id = grilla_formacion_ids[f_idx]
        for est_idx, minutos in enumerate(row):
            if minutos is None:
                continue
            conn.execute(
                "INSERT INTO horarios (grilla_formacion_id, grilla_estacion_id, minutos) VALUES (?, ?, ?)",
                (formacion_id, grilla_estacion_ids[est_idx], minutos),
            )
            horarios_insertados += 1
    return horarios_insertados


def replace_grid(conn, grilla_id: int, parsed: dict, station_ids: list[int]) -> int:
    delete_grid_content(conn, grilla_id)
    conn.execute(
        """
        UPDATE grillas
        SET hash_contenido = ?, metodo_actualizacion = ?, fuente_archivo = ?,
            actualizado_en = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (
            parsed["hash"], parsed["metodo_actualizacion"],
            parsed["fuente_archivo"], grilla_id,
        ),
    )
    return load_grid(conn, grilla_id, parsed, station_ids)


# ------------------------------------------------------------
# Procesamiento por archivo
# ------------------------------------------------------------

def process_file(conn, path: Path) -> tuple[str, str, dict]:
    stats_delta = {
        "ok": 0, "skip": 0, "rejected": 0,
        "grids_created": 0, "grids_updated": 0,
        "timetables_inserted": 0, "errors": 0,
    }

    try:
        parsed = parse_file(path)
        parsed["fuente_archivo"] = str(path)
    except Exception as exc:
        stats_delta["rejected"] += 1
        stats_delta["errors"] += 1
        log_import(conn, str(path), "error", str(exc))
        logger.error(f"Resultado: ERROR - Detalle: {exc}")
        return "ERROR", str(exc), stats_delta

    logger.info(f"Recorrido: {parsed['recorrido']}")
    logger.info(f"Día: {parsed['dia_canonical']}")
    logger.info(f"Sentido: {parsed['sentido_original']}")
    logger.info(f"Formaciones: {parsed['n']}")
    logger.info(f"Estaciones: {parsed['m']}")
    logger.info(f"Hash: {parsed['hash']}")

    try:
        conn.execute("BEGIN")
        entities = get_or_create_entities(conn, parsed)

        existing = conn.execute(
            """
            SELECT id, hash_contenido, metodo_actualizacion, fuente_archivo
            FROM grillas
            WHERE recorrido_id = ? AND dia_id = ? AND sentido_estacion_id = ?
            """,
            (entities["recorrido_id"], entities["dia_id"], entities["sentido_estacion_id"]),
        ).fetchone()

        if existing is None:
            cur = conn.execute(
                """
                INSERT INTO grillas (
                    recorrido_id, dia_id, sentido_estacion_id, hash_contenido,
                    metodo_actualizacion, fuente_archivo
                ) VALUES (?, ?, ?, ?, ?, ?)
                """,
                (entities["recorrido_id"], entities["dia_id"],
                 entities["sentido_estacion_id"], parsed["hash"],
                 parsed["metodo_actualizacion"], parsed["fuente_archivo"]),
            )
            grilla_id = cur.lastrowid
            inserted = load_grid(conn, grilla_id, parsed, entities["station_ids"])
            conn.commit()

            stats_delta["ok"] += 1
            stats_delta["grids_created"] += 1
            stats_delta["timetables_inserted"] += inserted

            detail = "Grilla creada"
            log_import(conn, str(path), "ok", detail)
            logger.info(f"Resultado: OK ({detail})")
            return "OK", detail, stats_delta

        if existing["hash_contenido"] == parsed["hash"]:
            grid_metadata_changed = (
                existing["metodo_actualizacion"] != parsed["metodo_actualizacion"]
                or existing["fuente_archivo"] != parsed["fuente_archivo"]
            )
            if grid_metadata_changed:
                conn.execute(
                    """
                    UPDATE grillas
                    SET metodo_actualizacion = ?, fuente_archivo = ?,
                        actualizado_en = CURRENT_TIMESTAMP
                    WHERE id = ?
                    """,
                    (
                        parsed["metodo_actualizacion"], parsed["fuente_archivo"],
                        existing["id"],
                    ),
                )
            conn.commit()
            stats_delta["skip"] += 1
            detail = (
                "Hash igual, metadata actualizada"
                if entities["source_changed"] or grid_metadata_changed
                else "Hash igual, sin cambios"
            )
            log_import(conn, str(path), "skip", detail)
            logger.info(f"Resultado: SKIP ({detail})")
            return "SKIP", detail, stats_delta

        inserted = replace_grid(conn, existing["id"], parsed, entities["station_ids"])
        conn.commit()

        stats_delta["ok"] += 1
        stats_delta["grids_updated"] += 1
        stats_delta["timetables_inserted"] += inserted

        detail = "Grilla reemplazada"
        log_import(conn, str(path), "ok", detail)
        logger.info(f"Resultado: OK ({detail})")
        return "OK", detail, stats_delta

    except Exception as exc:
        conn.rollback()
        stats_delta["rejected"] += 1
        stats_delta["errors"] += 1
        log_import(conn, str(path), "error", str(exc))
        logger.error(f"Resultado: ERROR - Detalle: {exc}")
        return "ERROR", str(exc), stats_delta


def discover_files(base_dir: Path = HORARIOS_DIR) -> list[Path]:
    if not base_dir.exists():
        return []
    files = list(base_dir.glob("*/*/*.xlsx"))
    return sorted(files, key=lambda p: str(p.resolve()))


def process_all_files(db_path=DB_PATH, base_dir: Path = HORARIOS_DIR) -> dict:
    conn = init_database(db_path)
    stats = new_stats()
    files = discover_files(base_dir)
    stats["found"] = len(files)
    logger.info(f"Archivos encontrados: {len(files)}")

    for path in files:
        logger.info(f"Procesando: {path}")
        estado, mensaje, delta = process_file(conn, path)
        for k, v in delta.items():
            if k in stats:
                stats[k] += v

    logger.info("Resumen:")
    logger.info(f"Archivos encontrados: {stats['found']}")
    logger.info(f"Archivos procesados correctamente: {stats['ok']}")
    logger.info(f"Archivos sin cambios: {stats['skip']}")
    logger.info(f"Archivos rechazados: {stats['rejected']}")
    logger.info(f"Grillas creadas: {stats['grids_created']}")
    logger.info(f"Grillas actualizadas: {stats['grids_updated']}")
    logger.info(f"Horarios insertados: {stats['timetables_inserted']}")
    logger.info(f"Errores: {stats['errors']}")

    conn.close()
    return stats


def rebuild_database_atomic(db_path: Path, base_dir: Path, strict: bool = True) -> dict:
    db_path = Path(db_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temp_name = tempfile.mkstemp(
        prefix=f".{db_path.name}.", suffix=".tmp", dir=db_path.parent,
    )
    os.close(descriptor)
    temp_path = Path(temp_name)
    temp_path.unlink(missing_ok=True)

    try:
        stats = process_all_files(temp_path, base_dir)
        if stats["found"] == 0:
            raise ScheduleImportError("No se encontraron libros XLSX")
        if strict and stats["errors"]:
            raise ScheduleImportError(
                f"La reconstrucción fue rechazada: {stats['errors']} libro(s) con error"
            )

        conn = sqlite3.connect(temp_path)
        try:
            integrity = conn.execute("PRAGMA integrity_check").fetchone()[0]
            grids = conn.execute("SELECT COUNT(*) FROM grillas").fetchone()[0]
            schedules = conn.execute("SELECT COUNT(*) FROM horarios").fetchone()[0]
        finally:
            conn.close()
        if integrity != "ok" or grids == 0 or schedules == 0:
            raise ScheduleImportError(
                f"Base inválida: integrity={integrity}, grillas={grids}, horarios={schedules}"
            )

        os.replace(temp_path, db_path)
        logger.info(
            "Base reemplazada atómicamente: %s (%s grillas, %s horarios)",
            db_path, grids, schedules,
        )
        return stats
    finally:
        temp_path.unlink(missing_ok=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--db", type=Path, default=DB_PATH)
    parser.add_argument("--horarios-dir", type=Path, default=HORARIOS_DIR)
    parser.add_argument("--rebuild", action="store_true")
    parser.add_argument("--strict", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        if args.rebuild:
            stats = rebuild_database_atomic(args.db, args.horarios_dir, args.strict)
        else:
            stats = process_all_files(args.db, args.horarios_dir)
    except ScheduleImportError as exc:
        logger.error("Resultado: ERROR - Detalle: %s", exc)
        return 1
    return 1 if args.strict and stats["errors"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
